import openpyxl
path = "/Users/riverwalser/Physics/wsj_headline_evals.xlsx"
import openpyxl
# To open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook(path, data_only=True)
 
# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active

#current date counter
for c in range(2, 19226):
    curr_d_object = sheet_obj.cell(row = c, column = 2)

    curr_date = str(curr_d_object.value).split(" ")[0]
    yday_object = sheet_obj.cell(row = c-1, column = 2)
    ydate = str(yday_object.value).split(" ")[0]
    print(f'current date: {curr_date}, ydate:{ydate}')
    if curr_date == ydate:
        sent_tdy = sheet_obj.cell(row=c,column=5).value
        print(sent_tdy)
        sent_yday = sheet_obj.cell(row=c-1,column=5).value
        print(sent_yday)
        avg_sent = (sent_tdy+sent_yday) / 2
        print(f'Average Sentiment is {avg_sent}')
        acc_tdy = sheet_obj.cell(row=c,column=6).value
        print(f'accuracy today: {acc_tdy}')
        acc_yday = sheet_obj.cell(row=c-1,column=6).value
        avg_acc = (acc_tdy + acc_yday) /2
        print(f'Average accuracy:{avg_acc}')
        tot_sent = sheet_obj.cell(row=c,column=7).value
        print(f'total sentiment today: {tot_sent}')
        sent_yday = sheet_obj.cell(row=c-1,column=7).value
        print(f'total sentiment yesterday:{sent_yday}')
        avg_sentiment_total = (tot_sent+ sent_yday)/2
        print(f'Average sentiment: {(tot_sent+sent_yday)/2}')
        sheet_obj.cell(row=c,column=5).value = avg_sent
        sheet_obj.cell(row=c,column=6).value = avg_acc
        sheet_obj.cell(row=c,column=7).value = avg_sentiment_total
        sheet_obj.cell(row=c-1,column=1).value = "DELET"
        print(f'working?: {sheet_obj.cell(row=c-1, column=1).value}')

wb_obj.save("/Users/riverwalser/Physics/verynewfiel.xlsx")
        
"""1477	5/4/98	Hongkong Telecom's Revenue Fell	Business and Finance - Asia	-1	0.999679804	-0.999679804
    3302	5/4/98	Boeing to Take Equity Stake in Ellipso	Major Business News	1	0.903332829	0.903332829"""
        
        
        
        
        
        