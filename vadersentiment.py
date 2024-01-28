# from vaderSentiment.vaderSentiment module.

 
# function to print sentiments
# of the sentence.
"""
def sentiment_scores(sentence):
    
 
    # Create a SentimentIntensityAnalyzer object.
    sid_obj = SentimentIntensityAnalyzer()
 
    # polarity_scores method of SentimentIntensityAnalyzer
    # object gives a sentiment dictionary.
    # which contains pos, neg, neu, and compound scores.
    sentiment_dict = sid_obj.polarity_scores(sentence)
     
    print("Overall sentiment dictionary is : ", sentiment_dict)
    print("sentence was rated as ", sentiment_dict['neg']*100, "% Negative")
    print("sentence was rated as ", sentiment_dict['neu']*100, "% Neutral")
    print("sentence was rated as ", sentiment_dict['pos']*100, "% Positive")
 
    print("Sentence Overall Rated As", end = " ")
 
    # decide sentiment as positive, negative and neutral
    if sentiment_dict['compound'] >= 0.05 :
        print("Positive")
   
    elif sentiment_dict['compound'] <= - 0.05 :
        print("Negative")
 
    else :
        print("Neutral")
        
    return [sentiment_dict['pos']*100, sentiment_dict['neu']*100, sentiment_dict['neg']*100]"""
    
from transformers import pipeline
import torch
def sentiment_scores(sentence):
    # Load the sentiment analysis model (BERT-based)
    classifier = pipeline("sentiment-analysis", model="distilbert-base-uncased-finetuned-sst-2-english")

    # Perform sentiment analysis on the sentence
    result = classifier(sentence)[0]

    # Print the sentiment result
    print("Label:", result["label"])
    print("Score:", result["score"])

    # Return the sentiment score
    return [result['label'], result["score"]]
        
sentiment_scores("Hongkong Telecom's Revenue Fell")

# Give the location of the file
path = "/Users/riverwalser/Physics/wsj_pre.xlsx"
import openpyxl
# To open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook(path)
 
# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active
 
# Cell objects also have a row, column,
# and coordinate attributes that provide
# location information for the cell.
 
# Note: The first row or
# column integer is 1, not 0.
 
# Cell object is created by using
# sheet object's cell(  m) method.



for i in range(1, 19227): 

  cell_obj = sheet_obj.cell(row = i, column = 3)
  analsys_obj_p = sheet_obj.cell(row = i, column=5)
  analsys_obj_s = sheet_obj.cell(row = i, column=6)

  
  # Print value of cell object
  # using the value attribute
  print(cell_obj.value)
  x = sentiment_scores(cell_obj.value)
  analsys_obj_p.value = x[0]
  analsys_obj_s.value = x[1]

  print(i)

wb_obj.save("/Users/riverwalser/Physics/wsj_headline_evals.xlsx")