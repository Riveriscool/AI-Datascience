import openpyxl

# Load the first Excel file
wb1 = openpyxl.load_workbook('/Users/riverwalser/Physics/AMZN (4).xlsx')
sheet1 = wb1.active

# Load the second Excel file
wb2 = openpyxl.load_workbook('/Users/riverwalser/Physics/weniii.xlsx')
sheet2 = wb2.active

# Create a new workbook for the integrated data
wb_integrated = openpyxl.Workbook()
sheet_integrated = wb_integrated.active

# Copy data from the first file to the integrated file
for row in sheet1.iter_rows(min_row=1, max_row=sheet1.max_row, values_only=True):
    date = row[0]  # Assuming the date is in column A
    row_data = row[:7]  # Copying data from column A to G

    # Search for a matching date in the second file
    for row2 in sheet2.iter_rows(min_row=1, max_row=sheet2.max_row, values_only=True):
        if row2[1] == date:  # Assuming the date is in column B of the second file
            row_data += row2[1:8]  # Copying data from column B to G
            break  # Stop searching after finding a match

    # Append the row data to the integrated file
    sheet_integrated.append(row_data)

# Save the integrated file
wb_integrated.save('integrated_file.xlsx')
